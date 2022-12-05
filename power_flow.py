#This file contains functions necessary to solve a power flow within an N bus system
#The program uses the newton raphson method to solve the power flow equations
#This class should be imported into a main file, which must provide an input data file and manage the looping of newton raphson iterations

import numpy as np
import pandas as pd
from openpyxl import Workbook
import math


class PowerFlow():
    #initialize the class with the specified number of maximum iterations to perform before failing and a
    #   specific tolerance the mismatches should be within for the solution to have converged
    def __init__(self, maxIterations, tolerance): 
        
        

        self.maxIterations = maxIterations #maximum number of iterations before failing
        self.currIterations = 0 #current number of iterations
        
        self.tolerance = tolerance #tolerance for solution to have converged

        self.BusData = 0 #input data regarding network buses
        self.LineData = 0 #input data regarding network lines
        self.numBusses = 0 #number of busses in network
        self.numConnections = 0 #number of line connections in network
        
        self.admittanceReal = 0 #real part of admittance matrix for network
        self.admittanceImag = 0 #imaginary part of admittance matrix for network
        
        self.PEquationList = 0 #list containing the current value of the active power flow equation for each node
        self.QEquationList = 0 #list containing the current value of the reactive power flow equation for each node
        self.implicitEquationList = 0 #list containing the current value of the mismatches for each implicit node
        
        self.deltaList = 0 #list containing the change in voltage and angle provided from the current newton raphson iteration
        self.inverseJacobian = 0 #the inverse jacobian of the matrix of power flow equations

        self.voltages = 0 #list of voltages at every node
        self.angles = 0 #list of angles at every node
        self.busType = 0 #list of the type of bus at each node
        self.numPQ = 0 #number of PQ buses
        self.numPV = 0 #number of PV buses

        self.busMap = 0 #list of original bus numbers, before sorting by program

        self.maxMismatchP = [] #list of the highest mismatch in active power at each NR iteration
        self.maxMismatchPLocations = [] #list of the locations for the above highest active power mismatches
        self.maxMismatchQ = [] #list of the highest mismatch in reactive power at each NR iteration
        self.maxMismatchQLocations = [] #list of the locations for the above highest reactive power mismatches

    #takes argument fileName, which is the name of the excel file to be read as input data
    #this function will pull voltages, load and generator power, line impedances, and maximum line flows from the input excel file
    #this function also calls several other functions to build the initial data structures needed for the NR method
    def readFromFile(self, fileName):
        self.BusData = pd.read_excel(fileName, sheet_name = 0)
        self.BusData = self.BusData.sort_values( by='Type', ascending = False) #sort data by bus type, so buses are in order: slack, pv, pq

        self.LineData = pd.read_excel(fileName, sheet_name = 1) 

        self.numBusses = len(self.BusData) # counts the number of buses
        self.numConnections = len(self.LineData) # counts the number of connections

        #make list to store original bus numbers before sorting, and indicate where they have moved after sorting
        self.busMap = np.array(self.BusData.index)

        #prepare the initial data structures that will not change over NR iterations
        self.buildAdmittanceMatrix()
        self.getVoltages()
        self.getBusType()

        #update equations matrix with updated initial data structures to prepare for 1st NR iteration
        self.updateEquationLists()

    #builds an admittance matrix based on the impedances loaded into the program through the readFromFile function
    def buildAdmittanceMatrix(self):
        
        admittanceComplex = np.zeros((self.numBusses, self.numBusses), dtype=np.complex_) #matrix storing complex impedance matrix
        self.admittanceReal = np.zeros((self.numBusses, self.numBusses)) #real part of impedance matrix
        self.admittanceImag = np.zeros((self.numBusses, self.numBusses)) #imaginary part of impedance matrix

        #iterate through each line connection, and for each add the corresponding value to the indexes of the admittance matrix it affects
        for i in range(self.numConnections):

            #these lines pull the from and to nodes listen in the file
            node1 = self.LineData['From'][i]
            node2 = self.LineData['To'][i]

            #these indexes store where the admittance for the specific node should go in the matrix
            node1Index = np.where(self.busMap == (node1 - 1))[0][0]
            node2Index = np.where(self.busMap == (node2 - 1))[0][0]           

            resistance1to2 = self.LineData['Rtotal, p.u.'][i]
            reactance1to2 = self.LineData['Xtotal, p.u.'][i]
            reactanceShunt = self.LineData['Btotal, p.u.'][i]

            admittance1to2 = 1/(resistance1to2 + 1j * reactance1to2) #add the admittance between nodes to corresponding off diagonal rows
            susceptance1and2 = (1j * reactanceShunt) / 2 #then add the admittance between nodes and 1/2 the susceptance to the diagonal rows

            #add in diagonal row values
            admittanceComplex[node1Index][node1Index] += ( admittance1to2 + susceptance1and2 )
            admittanceComplex[node2Index][node2Index] += ( admittance1to2 + susceptance1and2 )

            #add in non diagonal row values
            admittanceComplex[node1Index][node2Index] -= ( admittance1to2 )
            admittanceComplex[node2Index][node1Index] -= ( admittance1to2 )

        #get real and imaginary parts of complex matrix
        self.admittanceReal = np.real(admittanceComplex)
        self.admittanceImag = np.imag(admittanceComplex)

        print("Real Admittance:")
        print(self.admittanceReal)
        print()
        print("Imaginary Admittance:")
        print(self.admittanceImag)
        print()

    #gets voltages and angles from data pulled from input excel file    
    def getVoltages(self): 
        self.voltages = np.array(self.BusData['V Set'], dtype = float)
        self.angles = np.zeros_like(self.voltages, dtype = float)

        print("Voltages: ")
        print(self.voltages)
        print()

        print("Angles: ")
        print(self.angles)
        print()

    #gets type of bus from data pulled from input excel file. Type is either 'S' for slack, 'D' for PQ, 'G' for PV
    def getBusType(self):
        self.busType = np.array(self.BusData['Type'])
        self.numPQ = np.count_nonzero(self.busType == 'D')
        self.numPV = np.count_nonzero(self.busType == 'G')

    #this function will perform one iteration of the newton raphson method
    #The function will update the inverse jacobian matrix,
    #   update the required changes to voltage and angle at nodes based on that inverse jacobian matrix,
    #   add those updates to the corresponding voltages and angles,
    #   and finally update the power flow equations with these new values
    def newtonRaphsonIteration(self):
        #update all these values
        self.currIterations += 1
        
        self.updateInverseJacobian()
        self.updateDeltaList()
        self.updateVoltages()
        self.updateEquationLists()

        return self.implicitEquationList

    #this function updates the inverse jacobian matrix by dividing the jacobian into 4 regions, which are defined by certain equations
    #regions are H, M, N, and L which are built in separate functions
    def updateInverseJacobian(self):

        #jacobian will be of size N-1 + number of PQ buses
        jacobian = np.zeros( ( (self.numBusses-1) + self.numPQ, (self.numBusses-1) + self.numPQ), dtype = float) 

        jacobian = self.buildHMatrix(jacobian)
        jacobian = self.buildMMatrix(jacobian)
        jacobian = self.buildNMatrix(jacobian)
        jacobian = self.buildLMatrix(jacobian)

        # print("Jacobian:")
        # print(jacobian)
        print()
        self.inverseJacobian = np.linalg.inv(jacobian) #take inverse of jacobian matrix

    #builds the H portion of the jacobian matrix. Takes an empty jacobian matrix as an argument, and returns that matrix filled with the H portion
    #H is the top left portion of the jacobian matrix
    def buildHMatrix(self, jacobian):

        jacobianRow = 0 #the starting row of the jacobian for region H
        
        for k in range(2, self.numBusses + 1): #k goes from 2 to N
            k -= 1 #to use k for matrix indices that start at 0
            jacobianCol = 0 #the starting column of the jacobian for region H
            for i in range(2, self.numBusses + 1): #i goes from 2 to N
                i -= 1 #to use i for matrix indices that start at 0

                if i != k: #for non diagonal rows...
                    jacobian[jacobianRow][jacobianCol] = self.voltages[k]*self.voltages[i] * (
                            self.admittanceReal[k][i] * np.sin(self.angles[k] - self.angles[i]) -
                            self.admittanceImag[k][i] * np.cos(self.angles[k] - self.angles[i])
                        )

                else: #for diagonal rows...
                    for l in range(1, self.numBusses + 1):
                        l -= 1 #to use l for matrix indices that start at 0
                        if l != k:
                            jacobian[jacobianRow][jacobianCol] += self.voltages[k]*self.voltages[l] * (
                                    -1.0 * self.admittanceReal[k][l] * np.sin(self.angles[k] - self.angles[l]) +
                                    self.admittanceImag[k][l] * np.cos(self.angles[k] - self.angles[l])
                                )
                            
                jacobianCol += 1
            jacobianRow += 1
            
        return jacobian
    
    #builds the M portion of the jacobian matrix. Takes an empty jacobian matrix as an argument, and returns that matrix filled with the M portion
    #M is the top right portion of the jacobian matrix
    def buildMMatrix(self, jacobian):

        jacobianRow = 0 #the starting row of the jacobian for region M
        
        for k in range(2, self.numBusses + 1): #k goes from 2 to N
            k -= 1 #to use k for matrix indices that start at 0
            jacobianCol = self.numBusses - 1 #the starting column of the jacobian for region M
            for i in range((self.numPV + 1) + 1, self.numBusses + 1): #i goes from m+1 to N, m is number of PV buses + 1
                i -= 1 #to use i for matrix indices that start at 0

                if i != k: #for non diagonal rows
                    jacobian[jacobianRow][jacobianCol] = self.voltages[k] * (
                            self.admittanceReal[k][i] * np.cos(self.angles[k] - self.angles[i]) +
                            self.admittanceImag[k][i] * np.sin(self.angles[k] - self.angles[i])
                        )
                    
                    
                else: #for diagonal rows
                    jacobian[jacobianRow][jacobianCol] += 2 * self.admittanceReal[k][k] * self.voltages[k]
                    for l in range(1, self.numBusses + 1):
                        l -= 1 #to use l for matrix indices that start at 0
                        if l != k:
                            jacobian[jacobianRow][jacobianCol] += self.voltages[l] * (
                                    self.admittanceReal[k][l] * np.cos(self.angles[k] - self.angles[l]) +
                                    self.admittanceImag[k][l] * np.sin(self.angles[k] - self.angles[l])
                                )
                jacobianCol += 1
            jacobianRow += 1

        return jacobian
    
    #builds the N portion of the jacobian matrix. Takes an empty jacobian matrix as an argument, and returns that matrix filled with the N portion
    #N is the bottom left portion of the jacobian matrix
    def buildNMatrix(self, jacobian):

        jacobianRow = self.numBusses - 1 #the starting row of the jacobian for region N
        
        for k in range((self.numPV + 1) + 1, self.numBusses + 1): #k goes from m+1 to N, m is number of PV buses + 1
            k -= 1 #to use k for matrix indices that start at 0
            jacobianCol = 0 #the starting column of the jacobian for region N
            for i in range(2, self.numBusses + 1): #i goes 2 to N
                i -= 1 #to use i for matrix indices that start at 0

                if i != k: #for non diagonal rows
                    jacobian[jacobianRow][jacobianCol] = self.voltages[k]*self.voltages[i] * (
                            -1.0 * self.admittanceReal[k][i] * np.cos(self.angles[k] - self.angles[i]) +
                            -1.0 * self.admittanceImag[k][i] * np.sin(self.angles[k] - self.angles[i])
                        )
                    
                else: #for diagonal rows
                    for l in range(1, self.numBusses + 1):
                        l -= 1 #to use l for matrix indices that start at 0
                        if l != k:
                            jacobian[jacobianRow][jacobianCol] += self.voltages[k]*self.voltages[l] * (
                                    self.admittanceReal[k][l] * np.cos(self.angles[k] - self.angles[l]) +
                                    self.admittanceImag[k][l] * np.sin(self.angles[k] - self.angles[l])
                                )
                jacobianCol += 1
            jacobianRow += 1

        return jacobian
               
    #builds the L portion of the jacobian matrix. Takes an empty jacobian matrix as an argument, and returns that matrix filled with the L portion
    #L is the bottom right portion of the jacobian matrix
    def buildLMatrix(self, jacobian):

        jacobianRow = self.numBusses - 1 #the starting row of the jacobian for region L
        
        for k in range((self.numPV + 1) + 1, self.numBusses + 1): #k goes from m+1 to N, m is number of PV buses + 1
            k -= 1 #to use k for matrix indices that start at 0
            jacobianCol = self.numBusses - 1 #the starting column of the jacobian for region L
            for i in range((self.numPV + 1) + 1, self.numBusses + 1): #k goes from m+1 to N, m is number of PV buses + 1
                i -= 1 #to use i for matrix indices that start at 0

                if i != k: #for non diagonal rows
                    jacobian[jacobianRow][jacobianCol] = self.voltages[k] * (
                            self.admittanceReal[k][i] * np.sin(self.angles[k] - self.angles[i]) +
                            -1.0 * self.admittanceImag[k][i] * np.cos(self.angles[k] - self.angles[i])
                        )
                    
                else: #for diagonal rows
                    jacobian[jacobianRow][jacobianCol] += -2 * self.admittanceImag[k][k]*self.voltages[k]
                    for l in range(1, self.numBusses + 1):
                        l -= 1 #to use l for matrix indices that start at 0
                        if l != k:
                            jacobian[jacobianRow][jacobianCol] += self.voltages[l] * (
                                    self.admittanceReal[k][l] * np.sin(self.angles[k] - self.angles[l]) -
                                    self.admittanceImag[k][l] * np.cos(self.angles[k] - self.angles[l])
                                )
                jacobianCol += 1
            jacobianRow += 1

        return jacobian
                    
                
    #this function updates the required changes to voltages and angles indicated by the newton raphson method
    def updateDeltaList(self):
        print("Implicit Equation List:")
        print( np.vstack(self.implicitEquationList))
        print()
        
        self.deltaList = np.dot( (-1 * self.inverseJacobian), self.implicitEquationList) #Performs the basic function of the newton raphson method
        print("Delta List:")
        print(self.deltaList)
        print()

    #apply required changes to voltages and angles
    def updateVoltages(self):
        
        #delta list will have theta2 to thetaN, then vm+1 to vN

        angleIndex = 1 #index of first angle to be modified by delta list
        for i in range(self.numBusses - 1): #pull angles out of delta list and add them to angles list
            self.angles[angleIndex] += self.deltaList[i]
            angleIndex += 1

        voltageIndex = self.numPV + 1 #index of first voltage to be modified by delta list
        for i in range(self.numBusses - 1, len(self.deltaList)): #pull voltages out of delta list and add them to voltages list
            self.voltages[voltageIndex] += self.deltaList[i] 
            voltageIndex += 1

        print("Voltages:")
        print(self.voltages)
        print()

        print("Angles:")
        print(self.angles)
        print()

    #update power flow equations and implicit equations with the current values for voltage and angle at each bus
    def updateEquationLists(self):
        
        self.PEquationList = np.zeros(self.numBusses) #list of active power flow equation for each bus
        self.QEquationList = np.zeros(self.numBusses) #list of reactive power flow equation for each bus

        self.implicitEquationList = [] #list of implicit power flow equations, in order of of bus, with active equations before reactive
        #create P and Q equation list
        for i in range(self.numBusses):
            for j in range(self.numBusses):
                PSum = (self.voltages[i] * self.voltages[j]) * (
                            self.admittanceReal[i][j] * np.cos(self.angles[i] - self.angles[j]) +
                            self.admittanceImag[i][j] * np.sin(self.angles[i] - self.angles[j])
                        )

                QSum = (self.voltages[i] * self.voltages[j]) * (
                            self.admittanceReal[i][j] * np.sin(self.angles[i] - self.angles[j]) -
                            self.admittanceImag[i][j] * np.cos(self.angles[i] - self.angles[j])
                        )
                
                self.PEquationList[i] += PSum
                self.QEquationList[i] += QSum

        #for tracking the current highest active power mismatch and its location
        maximumPMismatch = 0 
        currPMismatch = 0 
        maximumPLocation = 0
            
        for i in range(len(self.PEquationList)): #get each item in the P list, see if it is implicit or not and add it if so
            if(self.busType[i] != 'S'): #both PQ and PV buses have an explicit P equation
                currPMismatch = self.PEquationList[i] - ((self.BusData['P Gen'][self.busMap[i]] / 100) - (self.BusData['P MW'][self.busMap[i]] / 100))
                if currPMismatch > maximumPMismatch: #if current mismatch is greater than maximum, set current to maximum
                    maximumPMismatch = currPMismatch
                    maximumPLocation = i
                    
                self.implicitEquationList.append([currPMismatch])

        #for tracking the current highest reactive power mismatch and its location
        maximumQMismatch = 0
        currQMismatch = 0
        maximumQLocation = 0

        for i in range(len(self.QEquationList)): #get each item in the Q list, see if it is implicit or not and add it if so
            if(self.busType[i] == 'D'): #only PQ buses have an explicit Q equation
                currQMismatch = self.QEquationList[i] + (self.BusData['Q MVAr'][self.busMap[i]] / 100)
                if currQMismatch > maximumQMismatch: #if current mismatch is greater than maximum, set current to maximum
                    maximumQMismatch = currQMismatch
                    maximumQLocation = i
                    
                self.implicitEquationList.append([currQMismatch])

        #if slack bus, none of the equations are explicit

        #build an array of maximum mismatches through all iterations
        self.maxMismatchP.append(maximumPMismatch)
        self.maxMismatchQ.append(maximumQMismatch)
        self.maxMismatchPLocations.append(maximumPLocation)
        self.maxMismatchQLocations.append(maximumQLocation)

        print("implicit equations...")
        print(np.vstack(self.implicitEquationList))

    #store as output excel file: voltages at each bus,
    #   power produced at each generator,
    #   power flow in each line connection,
    #   checks for violations of voltage and maximum power flow, and convergence record
    #
    #output excel file is indicated by argument filename
    def output(self, filename):

        #create workbook and sheets
        wb = Workbook()
        sheet1 = wb.create_sheet("List of Voltages", 0)
        sheet2 = wb.create_sheet("Power Produce at Gen.", 1)
        sheet3 = wb.create_sheet("Power Flow at Line", 2)
        sheet4 = wb.create_sheet("Line Check", 3)
        sheet5 = wb.create_sheet("Voltage Check", 4)
        sheet6 = wb.create_sheet("Convergence Record", 5)

        #build sheet1, list of voltages
        for i in range(self.numBusses):
            voltageStatement = "Bus ", i + 1, "Voltage: ", self.voltages[np.where(self.busMap == i)[0][0]], "Angle: ", self.angles[np.where(self.busMap == i)[0][0]] * (180 / math.pi)
            sheet1.append(voltageStatement)

        #build sheet2, power produced at each generator
        for i in range(len(self.PEquationList)):
            flowStatement = "Bus ", i + 1, "P MW: ", self.PEquationList[i] * 100, "Q MVAr: ", self.QEquationList[i] * 100 
            sheet2.append(flowStatement)

        #build sheet3 and sheet4, checks for violations of maximum power flow and the power flow at each line connection
        for i in range(self.numConnections):
            node1 = self.LineData['From'][i]
            node2 = self.LineData['To'][i]
            
            node1Index = np.where(self.busMap == node1 - 1)[0][0]
            node2Index = np.where(self.busMap == node2 - 1)[0][0]

            realPower = (self.voltages[node1Index] * self.voltages[node2Index]) * (
                            self.admittanceReal[node1Index][node2Index] * np.cos(self.angles[node1Index] - self.angles[node2Index]) + 
                            self.admittanceImag[node1Index][node2Index] * np.sin(self.angles[node1Index] - self.angles[node2Index])
                        )
            
            reactivePower = (self.voltages[node1Index] * self.voltages[np.where(self.busMap == node2 - 1)[0][0]]) * (
                                self.admittanceReal[node1Index][node2Index] * np.sin(self.angles[node1Index] - self.angles[node2Index]) -
                                self.admittanceImag[node1Index][node2Index] * np.cos(self.angles[node1Index] - self.angles[node2Index])
                            )
            MVA = math.sqrt((100 * realPower) ** 2 + (100 * reactivePower) ** 2)

            lineStatement = "Line:", node1, "to", node2, "P MW:", realPower * 100, "Q MVAr:", reactivePower * 100, "S MVA:", MVA
            sheet3.append(lineStatement)

            if  MVA > self.LineData['Fmax, MVA'][i]:
                lineData = "Line ", node1, " to Line ", node2, " has exceeded normal operation limit"
                sheet4.append(lineData)
                
            else:
                lineData = "Line ", node1, " to Line ", node2, " has not exceeded normal operation limit"
                sheet4.append(lineData)

        #build sheet5, checks for over/under voltage at each bus
        for i in range(self.numBusses):
            if self.voltages[np.where(self.busMap == i)[0][0]] > 1.05 or self.voltages[np.where(self.busMap == i)[0][0]] < 0.95:
                voltageData = "Bus ", i + 1, "has exceeded normal operating limits"
                sheet5.append(voltageData)
            else:
                voltageData = "Bus ", i + 1, "has not exceeded normal operating limits"
                sheet5.append(voltageData)

        #build sheet6, convergence record
        sheet6.append(["Iteration", "Maximum P Mismatch", "Location of P Mismatch", "Maximum Q Mismatch", "Location of Q Mismatch"])

        #trim extra mismatch data from mismatches before first newton raphson iteration
        self.maxMismatchP = self.maxMismatchP[1:len(self.maxMismatchP)]
        self.maxMismatchPLocations = self.maxMismatchPLocations[1:len(self.maxMismatchPLocations)]
        self.maxMismatchQ = self.maxMismatchQ[1:len(self.maxMismatchQ)]
        self.maxMismatchQLocations = self.maxMismatchQLocations[1:len(self.maxMismatchQLocations)]

        #add a row for each newton raphson iteration with the maximum mismatches
        for i in range(len(self.maxMismatchP)):
            sheet6.append([i + 1, self.maxMismatchP[i], self.maxMismatchPLocations[i], self.maxMismatchQ[i], self.maxMismatchQLocations[i]])

        wb.save(filename)
